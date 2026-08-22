# -*- coding: utf-8 -*-
"""Build the race audit workbook: every polling unit in a race, one row each.

  python3 scripts/build_audit_workbook.py \
      --state 29 --race "2026 Osun State Governorship" \
      --run storage/audit-osun2026/vlm_merged.jsonl \
      --out storage/audit-osun2026/osun-2026-governorship-audit.xlsx

WHY EVERY PU AND NOT EVERY SHEET. The register is the population; the sheets are
what INEC actually published. A unit with no sheet is not a blank row to skip -
it is the audit's first finding, and it only exists if the workbook is built
from the register and the sheets are joined ONTO it. Building it the other way
round makes missing units invisible.

The workbook is a working document: the last columns are for the human doing the
review, and nothing downstream reads them - they are for the audit trail.
"""
import argparse, json, os, sqlite3, sys
from collections import Counter, OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OSUN_2026_BALLOT = ['A', 'AA', 'AAC', 'ADC', 'ADP', 'APC', 'APGA', 'APM', 'APP',
                    'BP', 'NNPP', 'PRP', 'SDP', 'YPP', 'ZLP']
BOXES = [('registered', '#1 Registered'), ('accredited', '#2 Accredited'),
         ('ballotsIssued', '#3 Issued'), ('unusedBallots', '#4 Unused'),
         ('spoiled', '#5 Spoiled'), ('rejected', '#6 Rejected'),
         ('totalValid', '#7 Valid votes'), ('usedBallots', '#8 Used ballots')]
CHECKS = [('party_sum', 'Party sum = #7'), ('ballot_account', '#5+#6+#7 = #8'),
          ('ballot_stock', '#3-#4 = #8'), ('over_voting', 'Cast <= accredited'),
          ('accredited_vs_registered', '#2 <= #1'), ('valid_vs_used', '#7 <= #8'),
          ('registered_vs_issued', '#1 = #3'),
          # The officer's own TOTAL VALID VOTES row, read from the party-table
          # crop. A fourth independent statement of #7, in his handwriting.
          ('total_row', "Officer's total = #7")]
REVIEW_COLS = ['Reviewed by', 'Date reviewed', 'Finding confirmed?', 'Reviewer notes']

FONT = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='1F3B2C')
GRP_FILL = PatternFill('solid', fgColor='E8F0EA')
FILL = {
    'pass': PatternFill('solid', fgColor='D7EAD9'),
    'fail': PatternFill('solid', fgColor='F6C6C0'),
    'unknown': PatternFill('solid', fgColor='EFEFEF'),
    # A constraint SPENT deciding between two candidate readings. Not a pass -
    # it was used to choose the value, so it cannot also vouch for it - and not
    # a failure either. Its own colour so the distinction survives into the
    # deliverable instead of being flattened into green.
    'assumed': PatternFill('solid', fgColor='DCE5F5'),
    'publishable': PatternFill('solid', fgColor='C6E6C9'),
    'flagged': PatternFill('solid', fgColor='F3B5AC'),
    'review': PatternFill('solid', fgColor='FDE9BE'),
    'nosheet': PatternFill('solid', fgColor='D9D2E9'),
    'input': PatternFill('solid', fgColor='FFF9D6'),
}
THIN = Side(style='thin', color='C9D3CC')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_register(db_path, state):
    db = sqlite3.connect('file:%s?mode=ro' % db_path, uri=True)
    like = state + '-%'
    rows = db.execute(
        'SELECT pu_code,name,ward,lga,senatorial,federal_constituency,registered_voters '
        'FROM polling_units WHERE pu_code LIKE ? ORDER BY lga,ward,pu_code', (like,)).fetchall()
    db.close()
    return rows


def load_run(path):
    """One record per sheet, keyed by PU code."""
    out = {}
    if not path or not os.path.exists(path):
        return out
    with open(path, encoding='utf-8') as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            try:
                r = json.loads(line)
            except Exception:
                continue
            key = os.path.basename(r.get('file', '')).rsplit('.', 1)[0]
            if key:
                out[key] = r
    return out


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color='FFFFFF', size=10)
        cell.fill = HDR_FILL
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.border = BORDER


def build(args):
    reg = load_register(args.db, args.state)
    if not reg:
        print('ABORT: no polling units for state code %s' % args.state); sys.exit(1)
    run = load_run(args.run)
    ballot = args.ballot.split(',') if args.ballot else OSUN_2026_BALLOT

    reg_codes = {r[0] for r in reg}
    orphans = sorted(k for k in run if k not in reg_codes)

    # Triage tiers, if stage1_triage.mjs has been run. Optional on purpose: the
    # workbook must still build before triage exists, and a missing tier file
    # should degrade to "everything is Tier C" rather than crash the deliverable.
    tiers = {}
    in_sample = set()
    run_dir = os.path.dirname(os.path.abspath(args.run))

    def code_of(item):
        # Tier A holds no-sheet units keyed by PU code and sheets keyed by file;
        # both have to land on the register row for the same unit.
        c = item.get('pu') or (item.get('file') or '').replace('.jpg', '')
        return c.replace('/', '-') if c else None

    p = os.path.join(run_dir, 'tier_a.json')
    if os.path.exists(p):
        for item in json.load(open(p, encoding='utf-8')):
            c = code_of(item)
            if c:
                tiers[c] = ('A', item.get('why', ''))

    # SAMPLE MEMBERSHIP IS NOT A TIER. Tier says what to DO with a sheet;
    # the sample is a fact about how it was drawn. Recording the sample as a
    # tier meant a drawn sheet that also needed exhaustive review got filed
    # under A and vanished from the sample — dropping precisely the material
    # sheets and biasing the estimate downward. A sheet can be both, and if it
    # is drawn it stays drawn.
    p = os.path.join(run_dir, 'tier_b.json')
    if os.path.exists(p):
        for item in json.load(open(p, encoding='utf-8')):
            c = code_of(item)
            if not c:
                continue
            in_sample.add(c)
            tiers.setdefault(c, ('B', 'random sample'))

    wb = Workbook()

    # ---------------------------------------------------------------- Audit
    ws = wb.active
    ws.title = 'Audit'
    headers = (['LGA', 'Ward', 'PU code', 'Polling unit', 'Sheet published']
               + [h for _, h in BOXES]
               + ballot
               + ['Rows read twice', 'Rows disagreeing', 'Rows single-sourced',
                  'Boxes reconciled', 'How reconciled']
               + [h for _, h in CHECKS]
               # Triage tier goes NEXT TO the verdict, because they answer
               # different questions and a reviewer needs both: the verdict says
               # whether the sheet reconciles, the tier says whether anyone
               # should spend time on it. 1,111 flagged sheets sorted by file
               # name is not a work queue.
               + ['Verdict', 'Tier', 'Why this tier', 'In sample'] + REVIEW_COLS)
    ws.append(headers)
    style_header(ws, 1, len(headers))

    # Derived from the header row, not counted by hand. These were literal
    # arithmetic over the section lengths, which is fine until a column is
    # added in the middle - then the check colouring silently lands on the
    # wrong cells and the workbook looks right while saying something else.
    first_check_col = headers.index(CHECKS[0][1]) + 1
    verdict_col = headers.index('Verdict') + 1
    first_num_col = headers.index(BOXES[0][1]) + 1
    last_num_col = headers.index('Rows single-sourced') + 1

    counts = Counter()
    for pu_code, name, ward, lga, sen, fed, reg_voters in reg:
        r = run.get(pu_code)
        sheet = (r or {}).get('sheet') or {}
        verify = (r or {}).get('verify') or {}
        summary = verify.get('summary') or {}
        checks = {c['name']: c for c in (verify.get('checks') or [])}

        published = 'YES' if r else 'NO'
        counts['published' if r else 'missing'] += 1

        row = [lga, ward, pu_code, name, published]
        for key, _ in BOXES:
            v = sheet.get(key)
            row.append(v if isinstance(v, int) else None)

        # Party votes: only the resolved value the verification stack accepted.
        # A cell it could not resolve is left EMPTY rather than zero - a blank
        # says "not established", a 0 says "no votes", and conflating them is
        # how an audit publishes a number nobody read.
        rows_by_party = {}
        for rr in (verify.get('rows') or []):
            if rr.get('party'):
                rows_by_party[str(rr['party']).upper()] = rr
        for p in ballot:
            rr = rows_by_party.get(p)
            row.append(rr.get('value') if rr and rr.get('value') is not None else None)

        row.append(summary.get('agree') if r else None)
        row.append(summary.get('conflict') if r else None)
        # Single-sourced rows: read by one pass only, or observed empty by one
        # cell while the other was unreadable, or carried over a pass that
        # contradicted itself. Usable, weaker, and never enough for publishable.
        row.append(summary.get('single') if r else None)
        # PROVENANCE. A box the two passes disagreed on, decided by the sheet's
        # own equations rather than read off the paper. It is a sound inference
        # - two independent constraints had to agree, and 6,706 adversarial
        # trials produced no wrong choice - but it is still an inference, and a
        # reader checking a number against the image deserves to know which
        # ones will not be sitting there in ink.
        adj = (r or {}).get('adjudicated') or []
        row.append(len(adj) if r else None)
        row.append(', '.join('%s=%s (%s)' % (a.get('field'), a.get('chose'), a.get('spent'))
                             for a in adj) or None)
        for key, _ in CHECKS:
            row.append((checks.get(key) or {}).get('status') if r else None)
        verdict = summary.get('verdict') if r else 'no sheet'
        row.append(verdict)
        counts['verdict:' + str(verdict)] += 1
        tier = tiers.get(pu_code)
        row.append(tier[0] if tier else ('A' if not r else 'C'))
        row.append(tier[1] if tier else ('no sheet published' if not r else ''))
        # Kept separate from Tier so a sheet drawn into the sample stays in it
        # even when it also needs exhaustive review.
        row.append('YES' if pu_code in in_sample else '')
        counts['tier:' + str(tier[0] if tier else ('A' if not r else 'C'))] += 1
        row.extend([None] * len(REVIEW_COLS))
        ws.append(row)

        n = ws.max_row
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=n, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if first_num_col <= c <= last_num_col:
                cell.number_format = '#,##0'
        if not r:
            ws.cell(row=n, column=5).fill = FILL['nosheet']
        for i, (key, _) in enumerate(CHECKS):
            st = (checks.get(key) or {}).get('status')
            if st in FILL:
                ws.cell(row=n, column=first_check_col + i).fill = FILL[st]
        vc = ws.cell(row=n, column=verdict_col)
        if verdict in FILL:
            vc.fill = FILL[verdict]
        elif verdict == 'no sheet':
            vc.fill = FILL['nosheet']
        vc.font = Font(name=FONT, size=10, bold=True)
        # Only the REVIEWER's columns get the input shading. Tier and its reason
        # sit after the verdict but are computed, and shading them as fill-in
        # fields invites someone to overwrite the triage.
        for c in range(len(headers) - len(REVIEW_COLS) + 1, len(headers) + 1):
            ws.cell(row=n, column=c).fill = FILL['input']

    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(headers)), ws.max_row)
    widths = {1: 20, 2: 22, 3: 15, 4: 34, 5: 15}
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 12)
    ws.row_dimensions[1].height = 42
    last_row = ws.max_row

    # -------------------------------------------------------------- Summary
    sm = wb.create_sheet('Summary')
    sm['A1'] = args.race
    sm['A1'].font = Font(name=FONT, bold=True, size=14)
    sm['A2'] = 'Every figure below is a formula over the Audit tab, so it follows any filter or correction made there.'
    sm['A2'].font = Font(name=FONT, italic=True, size=9)

    vcol = get_column_letter(verdict_col)
    pub = 'E2:E%d' % last_row
    vrng = '%s2:%s%d' % (vcol, vcol, last_row)

    sm['A4'] = 'Coverage'; sm['A4'].font = Font(name=FONT, bold=True, size=11)
    cov = [('Polling units in the register', '=COUNTA(Audit!C2:C%d)' % last_row),
           ('EC8A published by INEC', '=COUNTIF(Audit!%s,"YES")' % pub),
           ('NO sheet published', '=COUNTIF(Audit!%s,"NO")' % pub),
           ('% of units with a sheet', '=IFERROR(B6/B5,0)')]
    r0 = 5
    for i, (lab, f) in enumerate(cov):
        sm.cell(row=r0 + i, column=1, value=lab).font = Font(name=FONT, size=10)
        sm.cell(row=r0 + i, column=2, value=f).font = Font(name=FONT, size=10)
        sm.cell(row=r0 + i, column=2).number_format = '0.0%' if 'ratio' in f or '%' in lab else '#,##0'

    sm['A10'] = 'Verdict'; sm['A10'].font = Font(name=FONT, bold=True, size=11)
    sm['A11'] = 'What it means'; sm['A11'].font = Font(name=FONT, italic=True, size=9)
    verdicts = [
        ('publishable', 'every number read twice and every check ran and passed'),
        ('flagged', 'a check FAILED - for a human, not a conclusion'),
        ('review', 'something could not be established, so nothing is asserted'),
        ('no sheet', 'INEC published no EC8A for this unit'),
    ]
    for i, (v, meaning) in enumerate(verdicts):
        rr = 12 + i
        sm.cell(row=rr, column=1, value=v).font = Font(name=FONT, size=10, bold=True)
        sm.cell(row=rr, column=2, value='=COUNTIF(Audit!%s,"%s")' % (vrng, v)).font = Font(name=FONT, size=10)
        sm.cell(row=rr, column=2).number_format = '#,##0'
        sm.cell(row=rr, column=3, value=meaning).font = Font(name=FONT, italic=True, size=9)
        if v in FILL:
            sm.cell(row=rr, column=1).fill = FILL[v]

    sm['A18'] = 'Checks'; sm['A18'].font = Font(name=FONT, bold=True, size=11)
    for j, lab in enumerate(['Check', 'pass', 'fail', 'unknown']):
        c = sm.cell(row=19, column=1 + j, value=lab)
        c.font = Font(name=FONT, bold=True, size=10)
    for i, (key, label) in enumerate(CHECKS):
        col = get_column_letter(first_check_col + i)
        rng = 'Audit!%s2:%s%d' % (col, col, last_row)
        rr = 20 + i
        sm.cell(row=rr, column=1, value=label).font = Font(name=FONT, size=10)
        for j, st in enumerate(['pass', 'fail', 'unknown']):
            c = sm.cell(row=rr, column=2 + j, value='=COUNTIF(%s,"%s")' % (rng, st))
            c.font = Font(name=FONT, size=10)
            c.number_format = '#,##0'
            c.fill = FILL[st]
    sm.cell(row=27, column=1, value='A "fail" is a sheet whose own numbers disagree. It is NOT a finding until a human has looked: '
                                    'a misread is far more common than a bad sheet.').font = Font(name=FONT, italic=True, size=9)
    for c, w in {1: 42, 2: 14, 3: 52, 4: 14}.items():
        sm.column_dimensions[get_column_letter(c)].width = w

    # --------------------------------------------------------------- By LGA
    bl = wb.create_sheet('By LGA')
    bl.append(['LGA', 'Units', 'Sheet published', 'No sheet', 'Publishable', 'Flagged', 'Review'])
    style_header(bl, 1, 7)
    for lga in sorted({r[3] for r in reg}):
        safe = str(lga).replace('"', '""')
        n = bl.max_row + 1
        bl.cell(row=n, column=1, value=lga)
        bl.cell(row=n, column=2, value='=COUNTIF(Audit!$A$2:$A$%d,$A%d)' % (last_row, n))
        bl.cell(row=n, column=3, value='=COUNTIFS(Audit!$A$2:$A$%d,$A%d,Audit!$E$2:$E$%d,"YES")' % (last_row, n, last_row))
        bl.cell(row=n, column=4, value='=COUNTIFS(Audit!$A$2:$A$%d,$A%d,Audit!$E$2:$E$%d,"NO")' % (last_row, n, last_row))
        for j, v in enumerate(['publishable', 'flagged', 'review']):
            bl.cell(row=n, column=5 + j,
                    value='=COUNTIFS(Audit!$A$2:$A$%d,$A%d,Audit!$%s$2:$%s$%d,"%s")'
                          % (last_row, n, vcol, vcol, last_row, v))
        for c in range(1, 8):
            cell = bl.cell(row=n, column=c)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            if c > 1:
                cell.number_format = '#,##0'
    bl.freeze_panes = 'A2'
    bl.auto_filter.ref = 'A1:G%d' % bl.max_row
    for c, w in {1: 26, 2: 10, 3: 16, 4: 11, 5: 13, 6: 11, 7: 11}.items():
        bl.column_dimensions[get_column_letter(c)].width = w

    # -------------------------------------------------------------- Read me
    rm = wb.create_sheet('Read me', 0)
    lines = [
        (args.race + ' - polling unit audit', 14, True),
        ('', 10, False),
        ('One row per polling unit ON THE REGISTER, not per sheet published. A unit with no EC8A is a', 10, False),
        ('finding in its own right, so it appears with "Sheet published = NO" rather than being omitted.', 10, False),
        ('', 10, False),
        ('YOU FILL IN the four shaded columns at the far right of the Audit tab:', 11, True),
        ('    Reviewed by · Date reviewed · Finding confirmed? · Reviewer notes', 10, False),
        ('Everything to their left is machine-read and should not be edited - correcting a value there', 10, False),
        ('would put the workbook out of step with the underlying transcription.', 10, False),
        ('', 10, False),
        ('Example of a completed review row:', 11, True),
    ]
    for i, (txt, size, bold) in enumerate(lines):
        c = rm.cell(row=1 + i, column=1, value=txt)
        c.font = Font(name=FONT, size=size, bold=bold)
    ex_row = len(lines) + 2
    for j, (h, v) in enumerate(zip(REVIEW_COLS, ['A. Adeyemi', '2026-08-22', 'Yes - over-voting confirmed',
                                                 'Accreditation box clear at 345; cast 347. Escalated.'])):
        hc = rm.cell(row=ex_row, column=1 + j, value=h)
        hc.font = Font(name=FONT, bold=True, size=10)
        vc = rm.cell(row=ex_row + 1, column=1 + j, value=v)
        vc.font = Font(name=FONT, size=10)
        vc.fill = FILL['input']
        vc.border = BORDER
    key_row = ex_row + 3
    rm.cell(row=key_row, column=1, value='Colour key').font = Font(name=FONT, bold=True, size=11)
    for i, (lab, fk) in enumerate([
            ('pass - the check ran and the numbers agree', 'pass'),
            ('fail - the check ran and the numbers disagree (for a human; NOT yet a finding)', 'fail'),
            ('unknown - the check could not run, usually an unreadable box. Never treated as a pass', 'unknown'),
            ('no sheet - INEC published no EC8A for this unit', 'nosheet'),
            ('for you to fill in', 'input')]):
        rr = key_row + 1 + i
        c = rm.cell(row=rr, column=1, value=lab)
        c.font = Font(name=FONT, size=10)
        c.fill = FILL[fk]
        c.border = BORDER
    note = key_row + 7
    for i, txt in enumerate([
            'How a figure is established: each party score is written twice on an EC8A, in figures and in words.',
            'A value appears here only where those two readings agree, or where exactly one of them was legible.',
            'Where they disagreed the cell is left BLANK - blank means "not established", never zero.',
            '',
            'Source: %s' % (os.path.basename(args.run) if args.run else 'register only, no transcription run supplied'),
            'Register: polling_units table, %d units with code %s-*' % (len(reg), args.state),
            'Generated by scripts/build_audit_workbook.py',
    ]):
        rm.cell(row=note + i, column=1, value=txt).font = Font(name=FONT, italic=True, size=9)
    rm.column_dimensions['A'].width = 104
    for col in 'BCD':
        rm.column_dimensions[col].width = 30

    wb.save(args.out)
    print('wrote %s' % args.out)
    print('  units %d · published %d · missing %d · orphan sheets %d'
          % (len(reg), counts['published'], counts['missing'], len(orphans)))
    for k in sorted(k for k in counts if k.startswith('verdict:')):
        print('  %-22s %d' % (k.split(':', 1)[1], counts[k]))
    if orphans:
        print('  NOTE: %d sheet(s) have no register entry: %s' % (len(orphans), ', '.join(orphans[:5])))


if __name__ == '__main__':
    ap = argparse.ArgumentParser()
    ap.add_argument('--state', default='29')
    ap.add_argument('--race', default='2026 Osun State Governorship Election')
    ap.add_argument('--run', default='')
    ap.add_argument('--out', required=True)
    ap.add_argument('--db', default='storage/hawkeye.db')
    ap.add_argument('--ballot', default='')
    build(ap.parse_args())
